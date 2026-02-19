"""
LTL Quotes Report Generator

Generates a weekly report matching the format:
- Rows: Customers (sorted by total volume descending)
- Columns grouped by week: Booked, Rated, Total Quotes, % Rated
- TOTAL row at top
- Excel output with formatting matching the screenshot
"""
import io
import os
import pandas as pd
from datetime import datetime
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from shared.drive_client import DriveClient

# Directory containing shared CSV files
SHARED_DIR = Path(__file__).parent.parent / 'shared'

# Optional: for Excel formatting
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def get_week_number_from_folder_name(folder_name: str) -> int:
    """Extract week number from folder name like 'W4925 Quotes' -> 49"""
    prefix = folder_name.split()[0]  # "W4925"
    week_num = int(prefix[1:3])  # Extract "49"
    return week_num


def get_year_from_folder_name(folder_name: str) -> int:
    """Extract year from folder name like 'W4925 Quotes' -> 2025"""
    prefix = folder_name.split()[0]  # "W4925"
    year_suffix = int(prefix[3:5])  # Extract "25"
    return 2000 + year_suffix


def get_all_2025_folders(client: DriveClient) -> list[dict]:
    """Get all week folders for 2025, sorted by week number."""
    folders = client.search_folders("Quotes")

    week_folders = []
    for f in folders:
        if f['name'].startswith('W') and 'Quotes' in f['name']:
            try:
                year = get_year_from_folder_name(f['name'])
                if year == 2025:
                    week_num = get_week_number_from_folder_name(f['name'])
                    week_folders.append({'id': f['id'], 'name': f['name'], 'week': week_num})
            except (ValueError, IndexError):
                continue

    return sorted(week_folders, key=lambda x: x['week'])


def get_available_weeks(client: DriveClient, max_weeks: int = 12) -> list[dict]:
    """
    Get all available week folders with labels, sorted chronologically.
    Returns list of dicts with: id, name, week, year, label
    """
    folders = client.search_folders("Quotes")

    week_folders = []
    for f in folders:
        if f['name'].startswith('W') and 'Quotes' in f['name']:
            try:
                week_num = get_week_number_from_folder_name(f['name'])
                year = get_year_from_folder_name(f['name'])
                label = f"W{week_num:02d}Y{year % 100:02d}"
                folder_info = {
                    'id': f['id'],
                    'name': f['name'],
                    'week': week_num,
                    'year': year,
                    'label': label
                }
                week_folders.append(folder_info)
            except (ValueError, IndexError):
                continue

    # Sort chronologically (oldest to newest), then take most recent max_weeks
    week_folders = sorted(week_folders, key=lambda x: (x['year'], x['week']), reverse=True)[:max_weeks]
    # Reverse to get chronological order (oldest first)
    return list(reversed(week_folders))


# Global cache for loaded CSV data (shared across report generators)
_csv_cache: dict[str, pd.DataFrame] = {}


def clear_csv_cache():
    """Clear the CSV data cache."""
    global _csv_cache
    _csv_cache = {}


def load_csvs_from_folder(client: DriveClient, folder_id: str, folder_name: str) -> pd.DataFrame:
    """Load all CSVs from a folder and combine into a single DataFrame.
    Uses in-memory caching to avoid re-downloading the same data for different reports.
    """
    global _csv_cache

    # Check cache first
    if folder_id in _csv_cache:
        print(f"  Using cached data for {folder_name} ({len(_csv_cache[folder_id]):,} quotes)")
        return _csv_cache[folder_id]

    files = client.list_files_in_folder(folder_id, file_type='csv')

    all_data = []
    for file in files:
        if file['name'].endswith('.csv'):
            content = client.download_file_content(file['id'])
            df = pd.read_csv(io.BytesIO(content), low_memory=False)
            df['source_file'] = file['name']
            all_data.append(df)

    if not all_data:
        _csv_cache[folder_id] = pd.DataFrame()
        return pd.DataFrame()

    combined = pd.concat(all_data, ignore_index=True)
    print(f"  Loaded {len(combined):,} quotes from {len(files)} files in {folder_name}")

    # Cache the result
    _csv_cache[folder_id] = combined
    return combined


def calculate_week_stats(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate stats per customer for a single week."""
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    df['booked'] = df['booked'].astype(str).str.lower() == 'true'
    df['is_rated'] = df['rate'].notna() & (df['rate'].astype(str).str.strip() != '')

    stats = df.groupby('customer').agg(
        booked=('booked', 'sum'),
        rated=('is_rated', 'sum'),
        total_quotes=('customer', 'count')
    ).reset_index()

    stats['pct_rated'] = (stats['rated'] / stats['total_quotes'] * 100).round(2)

    return stats


def generate_report(client: DriveClient, selected_weeks: list[dict] = None, num_weeks: int = 4) -> tuple[pd.DataFrame, list[str]]:
    """
    Generate the full pivot report for selected weeks.

    Args:
        client: DriveClient instance
        selected_weeks: List of week folder dicts (from get_available_weeks). If None, falls back to num_weeks.
        num_weeks: Number of most recent weeks (only used if selected_weeks is None)
    """
    if selected_weeks is None:
        # Fallback to old behavior for backwards compatibility
        selected_weeks = get_available_weeks(client, max_weeks=num_weeks)

    week_folders = selected_weeks
    week_labels = [f['label'] for f in week_folders]

    print(f"\nGenerating report for weeks: {week_labels}")

    all_customers = set()
    week_data = {}

    for folder in week_folders:
        print(f"\nProcessing {folder['name']}...")
        df = load_csvs_from_folder(client, folder['id'], folder['name'])
        stats = calculate_week_stats(df)
        week_data[folder['label']] = stats
        all_customers.update(stats['customer'].tolist())

    # Build report rows
    report_rows = []
    for customer in all_customers:
        row = {'Customers': customer, '_total_volume': 0}

        for week_label in week_labels:
            stats = week_data[week_label]
            cust_stats = stats[stats['customer'] == customer]

            if not cust_stats.empty:
                booked = int(cust_stats['booked'].values[0])
                rated = int(cust_stats['rated'].values[0])
                total = int(cust_stats['total_quotes'].values[0])
                pct = cust_stats['pct_rated'].values[0]
            else:
                booked, rated, total, pct = 0, 0, 0, 0.0

            row[f'{week_label}_Booked'] = booked
            row[f'{week_label}_Rated'] = rated
            row[f'{week_label}_Total Quotes'] = total
            row[f'{week_label}_% Rated'] = pct
            row['_total_volume'] += total

        report_rows.append(row)

    report_df = pd.DataFrame(report_rows)

    # Sort by total volume descending
    report_df = report_df.sort_values('_total_volume', ascending=False)
    report_df = report_df.drop(columns=['_total_volume'])

    # Calculate TOTAL row
    total_row = {'Customers': 'TOTAL'}

    for week_label in week_labels:
        total_row[f'{week_label}_Booked'] = report_df[f'{week_label}_Booked'].sum()
        total_row[f'{week_label}_Rated'] = report_df[f'{week_label}_Rated'].sum()
        total_row[f'{week_label}_Total Quotes'] = report_df[f'{week_label}_Total Quotes'].sum()
        total_quotes = total_row[f'{week_label}_Total Quotes']
        if total_quotes > 0:
            total_row[f'{week_label}_% Rated'] = round(
                total_row[f'{week_label}_Rated'] / total_quotes * 100, 2
            )
        else:
            total_row[f'{week_label}_% Rated'] = 0.0

    report_df = pd.concat([pd.DataFrame([total_row]), report_df], ignore_index=True)

    return report_df, week_labels


def get_month_from_week(week_num: int, year: int) -> tuple[int, int, str]:
    """
    Convert ISO week number to month.
    Returns (year, month_num, label) e.g., (2025, 1, "Jan 2025")
    """
    from datetime import datetime
    # Use the Thursday of the ISO week to determine the month
    # (ISO weeks start on Monday, and Thursday determines which year the week belongs to)
    date = datetime.strptime(f'{year}-W{week_num:02d}-4', "%G-W%V-%u")
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    return (date.year, date.month, f"{month_names[date.month-1]} {date.year}")


def aggregate_weeks_to_months(week_folders: list[dict]) -> dict[str, list[dict]]:
    """
    Group week folders by month.
    Returns dict mapping month_label -> list of week folder dicts
    """
    months = {}
    for folder in week_folders:
        year, month_num, month_label = get_month_from_week(folder['week'], folder['year'])
        if month_label not in months:
            months[month_label] = {'folders': [], 'year': year, 'month': month_num}
        months[month_label]['folders'].append(folder)

    # Sort by year and month
    sorted_labels = sorted(months.keys(),
                          key=lambda x: (months[x]['year'], months[x]['month']))
    return {label: months[label]['folders'] for label in sorted_labels}


def generate_monthly_report(client: DriveClient, selected_weeks: list[dict]) -> tuple[pd.DataFrame, list[str]]:
    """
    Generate report aggregated by month instead of by week.
    Used when more than 12 weeks are selected.
    """
    week_folders = selected_weeks

    # Group weeks by month
    months_dict = aggregate_weeks_to_months(week_folders)
    month_labels = list(months_dict.keys())

    print(f"\nGenerating MONTHLY report for {len(week_folders)} weeks grouped into {len(month_labels)} months")
    print(f"Months: {month_labels}")

    all_customers = set()
    month_data = {}  # month_label -> {customer -> {booked, rated, total}}

    # Load all week data first
    for folder in week_folders:
        print(f"\nProcessing {folder['name']}...")
        df = load_csvs_from_folder(client, folder['id'], folder['name'])
        stats = calculate_week_stats(df)

        # Find which month this week belongs to
        _, _, month_label = get_month_from_week(folder['week'], folder['year'])

        if month_label not in month_data:
            month_data[month_label] = {}

        # Aggregate into month
        for _, row in stats.iterrows():
            customer = row['customer']
            all_customers.add(customer)
            if customer not in month_data[month_label]:
                month_data[month_label][customer] = {'booked': 0, 'rated': 0, 'total': 0}
            month_data[month_label][customer]['booked'] += int(row['booked'])
            month_data[month_label][customer]['rated'] += int(row['rated'])
            month_data[month_label][customer]['total'] += int(row['total_quotes'])

    # Build report rows
    report_rows = []
    for customer in all_customers:
        row = {'Customers': customer, '_total_volume': 0}

        for month_label in month_labels:
            cust_data = month_data.get(month_label, {}).get(customer, {'booked': 0, 'rated': 0, 'total': 0})
            booked = cust_data['booked']
            rated = cust_data['rated']
            total = cust_data['total']
            pct = (rated / total * 100) if total > 0 else 0.0

            row[f'{month_label}_Booked'] = booked
            row[f'{month_label}_Rated'] = rated
            row[f'{month_label}_Total Quotes'] = total
            row[f'{month_label}_% Rated'] = round(pct, 2)
            row['_total_volume'] += total

        report_rows.append(row)

    report_df = pd.DataFrame(report_rows)

    # Handle empty report
    if report_df.empty or len(all_customers) == 0:
        # Return empty dataframe with proper structure
        empty_cols = ['Customers']
        for month_label in month_labels:
            empty_cols.extend([f'{month_label}_Booked', f'{month_label}_Rated',
                              f'{month_label}_Total Quotes', f'{month_label}_% Rated'])
        return pd.DataFrame(columns=empty_cols), month_labels

    # Sort by total volume descending
    if '_total_volume' in report_df.columns:
        report_df = report_df.sort_values('_total_volume', ascending=False)
        report_df = report_df.drop(columns=['_total_volume'])

    # Calculate TOTAL row
    total_row = {'Customers': 'TOTAL'}
    for month_label in month_labels:
        booked_col = f'{month_label}_Booked'
        rated_col = f'{month_label}_Rated'
        total_col = f'{month_label}_Total Quotes'
        pct_col = f'{month_label}_% Rated'

        if booked_col in report_df.columns:
            total_row[booked_col] = report_df[booked_col].sum()
            total_row[rated_col] = report_df[rated_col].sum()
            total_row[total_col] = report_df[total_col].sum()
            total_quotes = total_row[total_col]
            if total_quotes > 0:
                total_row[pct_col] = round(total_row[rated_col] / total_quotes * 100, 2)
            else:
                total_row[pct_col] = 0.0
        else:
            total_row[booked_col] = 0
            total_row[rated_col] = 0
            total_row[total_col] = 0
            total_row[pct_col] = 0.0

    report_df = pd.concat([pd.DataFrame([total_row]), report_df], ignore_index=True)

    return report_df, month_labels



def save_to_excel(report_df: pd.DataFrame, weeks: list[int], filename: str):
    """Save report to Excel with formatting matching the screenshot."""
    if not HAS_OPENPYXL:
        print("openpyxl not installed, saving as CSV instead")
        report_df.to_csv(filename.replace('.xlsx', '.csv'), index=False)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotes Report"

    # Colors
    header_green = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    black_font_bold = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Alternating week colors (for columns)
    week_colors = [light_green, white_fill]  # Alternates: green, white, green, white...
    week_header_colors = [light_green, light_gray]  # For header row

    # Row 1: Title
    total_cols = 1 + len(weeks) * 4
    ws.merge_cells('A1:' + openpyxl.utils.get_column_letter(total_cols) + '1')
    ws['A1'] = "WARP FREIGHT QUOTES RETURNED"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = header_green
    ws['A1'].alignment = Alignment(horizontal='center')

    # Row 2: Week numbers header
    ws['A2'] = ""
    ws['A2'].fill = header_green
    col = 2
    for week in weeks:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+3)
        cell = ws.cell(row=2, column=col, value=str(week))
        cell.font = white_font
        cell.fill = header_green
        cell.alignment = Alignment(horizontal='center')
        col += 4

    # Row 3: Sub-headers (Booked, Rated, Total Quotes, % Rated) with alternating colors
    ws['A3'] = "Customers"
    ws['A3'].font = black_font_bold
    ws['A3'].fill = light_green
    ws['A3'].border = thin_border
    col = 2
    for i, week in enumerate(weeks):
        header_fill = week_header_colors[i % 2]
        for header in ['Booked', 'Rated', 'Total Quotes', '% Rated']:
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = black_font_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            col += 1

    # Data rows with alternating week column colors
    for row_idx, row_data in enumerate(report_df.values, start=4):
        customer = row_data[0]
        cust_cell = ws.cell(row=row_idx, column=1, value=customer)
        cust_cell.border = thin_border

        col = 2
        for i, week in enumerate(weeks):
            base_idx = 1 + i * 4
            booked = row_data[base_idx]
            rated = row_data[base_idx + 1]
            total = row_data[base_idx + 2]
            pct = row_data[base_idx + 3]

            # Alternating fill for week columns
            week_fill = week_colors[i % 2]

            for j, val in enumerate([booked, rated, total]):
                cell = ws.cell(row=row_idx, column=col+j, value=val)
                cell.border = thin_border
                cell.fill = week_fill
                cell.number_format = '#,##0'  # Add thousands separator

            pct_cell = ws.cell(row=row_idx, column=col+3, value=f"{pct:.2f}%")
            pct_cell.border = thin_border
            pct_cell.alignment = Alignment(horizontal='right')
            pct_cell.fill = week_fill
            col += 4

        # Highlight TOTAL row (keep it green for emphasis)
        if customer == 'TOTAL':
            for c in range(1, total_cols + 1):
                ws.cell(row=row_idx, column=c).font = black_font_bold
                ws.cell(row=row_idx, column=c).fill = light_green

    # Adjust column widths
    ws.column_dimensions['A'].width = 45
    for col in range(2, total_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 13

    wb.save(filename)
    print(f"✓ Excel report saved to {filename}")


def load_zip_to_airport_mapping() -> dict:
    """Load zip code to airport code mapping from CSV file."""
    # Path to shared CSV file
    csv_path = SHARED_DIR / 'zip_to_airport_code - Sheet1.csv'

    if csv_path.exists():
        df = pd.read_csv(csv_path, dtype={'Zip Code': str})
        # Create mapping dict: zip_code -> airport_code
        mapping = dict(zip(df['Zip Code'].astype(str).str.zfill(5), df['Airport Code']))
        print(f"  Loaded {len(mapping):,} zip-to-airport mappings")
        return mapping

    print(f"  Warning: Could not find zip_to_airport_code mapping file at {csv_path}")
    return {}


def get_airport_code(zip_code: str, mapping: dict) -> str:
    """Get airport code for a zip code, or return the zip itself if not mapped."""
    if pd.isna(zip_code) or str(zip_code).strip() == '':
        return 'UNKNOWN'

    # Normalize zip code to 5 digits
    zip_str = str(zip_code).split('-')[0].split('.')[0].strip().zfill(5)[:5]

    return mapping.get(zip_str, zip_str)


def calculate_lanes_stats(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """Calculate stats per lane (origin-destination airport pair) for a single week.
    Only counts quotes that were rated (have a value in 'rate' column)."""
    if df.empty:
        return pd.DataFrame()

    df = df.copy()

    # Only include rated quotes
    df['is_rated'] = df['rate'].notna() & (df['rate'].astype(str).str.strip() != '')
    rated_df = df[df['is_rated']].copy()

    if rated_df.empty:
        return pd.DataFrame()

    # Map zip codes to airport codes
    rated_df['origin_airport'] = rated_df['pickup Zip'].apply(lambda x: get_airport_code(x, mapping))
    rated_df['dest_airport'] = rated_df['dropoff Zip'].apply(lambda x: get_airport_code(x, mapping))

    # Create lane identifier
    rated_df['lane'] = rated_df['origin_airport'] + '-' + rated_df['dest_airport']

    # Count quotes per lane
    stats = rated_df.groupby('lane').agg(
        total=('lane', 'count')
    ).reset_index()

    return stats


def generate_lanes_report(client: DriveClient, selected_weeks: list[dict] = None, num_weeks: int = 4) -> tuple[pd.DataFrame, list[str]]:
    """
    Generate the lanes quoted report for selected weeks.

    Args:
        client: DriveClient instance
        selected_weeks: List of week folder dicts (from get_available_weeks). If None, falls back to num_weeks.
        num_weeks: Number of most recent weeks (only used if selected_weeks is None)
    """
    # Load zip to airport mapping
    mapping = load_zip_to_airport_mapping()

    if selected_weeks is None:
        selected_weeks = get_available_weeks(client, max_weeks=num_weeks)

    week_folders = selected_weeks
    week_labels = [f['label'] for f in week_folders]

    print(f"\nGenerating lanes report for weeks: {week_labels}")

    all_lanes = set()
    week_data = {}

    for folder in week_folders:
        print(f"\nProcessing lanes for {folder['name']}...")
        df = load_csvs_from_folder(client, folder['id'], folder['name'])
        stats = calculate_lanes_stats(df, mapping)
        week_data[folder['label']] = stats
        if not stats.empty:
            all_lanes.update(stats['lane'].tolist())

    # Build report rows
    report_rows = []
    for lane in all_lanes:
        row = {'Lanes': lane}
        latest_week_label = week_labels[-1] if week_labels else None

        for i, week_label in enumerate(week_labels):
            stats = week_data[week_label]
            lane_stats = stats[stats['lane'] == lane] if not stats.empty else pd.DataFrame()

            if not lane_stats.empty:
                total = int(lane_stats['total'].values[0])
            else:
                total = 0

            row[f'{week_label}_Total'] = total

            # Calculate % change from previous week
            if i > 0:
                prev_week_label = week_labels[i - 1]
                prev_total = row.get(f'{prev_week_label}_Total', 0)
                if prev_total > 0:
                    pct_change = ((total - prev_total) / prev_total) * 100
                    row[f'{week_label}_%Change'] = pct_change
                else:
                    row[f'{week_label}_%Change'] = None  # Can't calculate % change from 0
            else:
                row[f'{week_label}_%Change'] = None  # No previous week to compare

        # Store latest week total for sorting
        row['_latest_total'] = row.get(f'{latest_week_label}_Total', 0) if latest_week_label else 0
        report_rows.append(row)

    report_df = pd.DataFrame(report_rows)

    if report_df.empty:
        return pd.DataFrame(), week_labels

    # Sort by latest week total volume descending
    report_df = report_df.sort_values('_latest_total', ascending=False)
    report_df = report_df.drop(columns=['_latest_total'])

    return report_df, week_labels


def load_airport_to_region_mapping() -> dict:
    """Load airport code to region mapping from CSV file."""
    # Path to shared CSV file
    csv_path = SHARED_DIR / 'airport_code_to_region - Sheet1.csv'

    if csv_path.exists():
        df = pd.read_csv(csv_path)
        # Filter out empty rows
        df = df.dropna(subset=['Airport Code', 'Region'])
        df = df[df['Airport Code'].str.strip() != '']
        # Create mapping dict: airport_code -> region
        mapping = dict(zip(df['Airport Code'].str.strip(), df['Region'].str.strip()))
        print(f"  Loaded {len(mapping):,} airport-to-region mappings")
        return mapping

    print(f"  Warning: Could not find airport_code_to_region mapping file at {csv_path}")
    return {}


def get_region(airport_code: str, mapping: dict) -> str:
    """Get region for an airport code, or return the airport code itself if not mapped."""
    if pd.isna(airport_code) or str(airport_code).strip() == '':
        return 'UNKNOWN'

    airport_str = str(airport_code).strip().upper()
    return mapping.get(airport_str, airport_str)


def calculate_regions_stats(df: pd.DataFrame, zip_mapping: dict, region_mapping: dict) -> pd.DataFrame:
    """Calculate stats per region pair (origin-destination region) for a single week.
    Only counts quotes that were rated (have a value in 'rate' column)."""
    if df.empty:
        return pd.DataFrame()

    df = df.copy()

    # Only include rated quotes
    df['is_rated'] = df['rate'].notna() & (df['rate'].astype(str).str.strip() != '')
    rated_df = df[df['is_rated']].copy()

    if rated_df.empty:
        return pd.DataFrame()

    # Map zip codes to airport codes, then airport codes to regions
    rated_df['origin_airport'] = rated_df['pickup Zip'].apply(lambda x: get_airport_code(x, zip_mapping))
    rated_df['dest_airport'] = rated_df['dropoff Zip'].apply(lambda x: get_airport_code(x, zip_mapping))

    rated_df['origin_region'] = rated_df['origin_airport'].apply(lambda x: get_region(x, region_mapping))
    rated_df['dest_region'] = rated_df['dest_airport'].apply(lambda x: get_region(x, region_mapping))

    # Create region lane identifier
    rated_df['region_lane'] = rated_df['origin_region'] + '-' + rated_df['dest_region']

    # Count quotes per region lane
    stats = rated_df.groupby('region_lane').agg(
        total=('region_lane', 'count')
    ).reset_index()

    return stats


def generate_regions_report(client: DriveClient, selected_weeks: list[dict] = None, num_weeks: int = 4) -> tuple[pd.DataFrame, list[str]]:
    """
    Generate the regions quoted report for selected weeks.

    Args:
        client: DriveClient instance
        selected_weeks: List of week folder dicts (from get_available_weeks). If None, falls back to num_weeks.
        num_weeks: Number of most recent weeks (only used if selected_weeks is None)
    """
    # Load mappings
    zip_mapping = load_zip_to_airport_mapping()
    region_mapping = load_airport_to_region_mapping()

    if selected_weeks is None:
        selected_weeks = get_available_weeks(client, max_weeks=num_weeks)

    week_folders = selected_weeks
    week_labels = [f['label'] for f in week_folders]

    print(f"\nGenerating regions report for weeks: {week_labels}")

    all_region_lanes = set()
    week_data = {}

    for folder in week_folders:
        print(f"\nProcessing regions for {folder['name']}...")
        df = load_csvs_from_folder(client, folder['id'], folder['name'])
        stats = calculate_regions_stats(df, zip_mapping, region_mapping)
        week_data[folder['label']] = stats
        if not stats.empty:
            all_region_lanes.update(stats['region_lane'].tolist())

    # Build report rows
    report_rows = []
    for region_lane in all_region_lanes:
        row = {'Regions': region_lane}
        latest_week_label = week_labels[-1] if week_labels else None

        for i, week_label in enumerate(week_labels):
            stats = week_data[week_label]
            lane_stats = stats[stats['region_lane'] == region_lane] if not stats.empty else pd.DataFrame()

            if not lane_stats.empty:
                total = int(lane_stats['total'].values[0])
            else:
                total = 0

            row[f'{week_label}_Total'] = total

            # Calculate % change from previous week
            if i > 0:
                prev_week_label = week_labels[i - 1]
                prev_total = row.get(f'{prev_week_label}_Total', 0)
                if prev_total > 0:
                    pct_change = ((total - prev_total) / prev_total) * 100
                    row[f'{week_label}_%Change'] = pct_change
                else:
                    row[f'{week_label}_%Change'] = None  # Can't calculate % change from 0
            else:
                row[f'{week_label}_%Change'] = None  # No previous week to compare

        # Store latest week total for sorting
        row['_latest_total'] = row.get(f'{latest_week_label}_Total', 0) if latest_week_label else 0
        report_rows.append(row)

    report_df = pd.DataFrame(report_rows)

    if report_df.empty:
        return pd.DataFrame(), week_labels

    # Sort by latest week total volume descending
    report_df = report_df.sort_values('_latest_total', ascending=False)
    report_df = report_df.drop(columns=['_latest_total'])

    return report_df, week_labels


if __name__ == "__main__":
    client = DriveClient()

    print("\n" + "="*60)
    print("WARP FREIGHT QUOTES RETURNED - Weekly Report")
    print("="*60)

    report, weeks = generate_report(client, num_weeks=4)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Save to CSV (simple format)
    csv_file = f"ltl_report_{timestamp}.csv"
    report.to_csv(csv_file, index=False)
    print(f"\n✓ CSV report saved to {csv_file}")

    # Save to Excel with formatting
    xlsx_file = f"ltl_report_{timestamp}.xlsx"
    save_to_excel(report, weeks, xlsx_file)

    # Display preview
    print("\nReport Preview (first 10 rows):")
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    print(report.head(10).to_string())
